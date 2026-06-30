from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Employee, Department, Designation, LeaveRequest, Payroll, Attendance
from .forms import EmployeeForm, DepartmentForm, DesignationForm, LeaveRequestForm, PayrollForm, AttendanceForm

@login_required(login_url='login')
def employee_list_view(request):
    employees = Employee.objects.all().order_by('-created_at')
    context={
        'employees':employees
    }
    return render(request, 'hrm/employee_list.html', context)


@login_required(login_url='login')
def employee_create_view(request): 

    if request.method == 'POST': 
        form = EmployeeForm(request.POST) 
        if form.is_valid(): 
            form.save() 
            return redirect('hrm:employee_list') 
    else: 
        form = EmployeeForm() 
    return render(request, 'hrm/employee_form.html', {'form': form}) 

@login_required(login_url='login')
def employee_edit_view(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('hrm:employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'hrm/employee_form.html', {'form': form})

@login_required(login_url='login')
def employee_delete_view(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('hrm:employee_list')
    return render(request, 'hrm/employee_confirm_delete.html', {'employee': employee})

@login_required(login_url='login')
def department_list_view(request):
    departments = Department.objects.all().order_by('-created_at')
    context={
        'departments':departments
    }
    return render(request, 'hrm/department_list.html', context)

@login_required(login_url='login')
def department_create_view(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('hrm:department_list')
    else:
        form = DepartmentForm()
    return render(request, 'hrm/department_form.html', {'form': form})



@login_required(login_url='login')
def designation_list_view(request):
    designations = Designation.objects.all().order_by('-created_at')
    return render(request, 'hrm/designation_list.html', {'designations': designations})

@login_required(login_url='login')
def designation_create_view(request):
    if request.method == 'POST':
        form = DesignationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('hrm:designation_list')
    else:
        form = DesignationForm()
    return render(request, 'hrm/designation_form.html', {'form': form})

@login_required(login_url='login')
def leave_list_view(request):
    leaves = LeaveRequest.objects.all().order_by('-created_at')
    return render(request, 'hrm/leave_list.html', {'leaves': leaves})

@login_required(login_url='login')
def leave_create_view(request):
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('hrm:leave_list')
    else:
        form = LeaveRequestForm()
    return render(request, 'hrm/leave_form.html', {'form': form})

@login_required(login_url='login')
def payroll_list_view(request):
    payrolls = Payroll.objects.all().order_by('-created_at')
    return render(request, 'hrm/payroll_list.html', {'payrolls': payrolls})

@login_required(login_url='login')
def payroll_create_view(request):
    if request.method == 'POST':
        form = PayrollForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('hrm:payroll_list')
    else:
        form = PayrollForm()
    return render(request, 'hrm/payroll_form.html', {'form': form})

@login_required(login_url='login')
def attendance_list_view(request):
    attendances = Attendance.objects.all().order_by('-date')
    return render(request, 'hrm/attendance_list.html', {'attendances': attendances})

@login_required(login_url='login')
def attendance_create_view(request):
    if request.method == 'POST':
        form = AttendanceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('hrm:attendance_list')
    else:
        form = AttendanceForm()
    return render(request, 'hrm/attendance_form.html', {'form': form})


@login_required(login_url='login')
def attendance_report_view(request):
    """Attendance report with date range filter."""
    from datetime import date
    today = date.today()

    # Default to current month
    date_from_str = request.GET.get('date_from', today.replace(day=1).strftime('%Y-%m-%d'))
    date_to_str   = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    try:
        from datetime import datetime
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        date_from = today.replace(day=1)
        date_to   = today

    attendances = (
        Attendance.objects
        .filter(date__range=[date_from, date_to])
        .select_related('employee', 'employee__department', 'employee__designation')
        .order_by('date', 'employee__first_name')
    )

    # Summary stats
    total   = attendances.count()
    present = attendances.filter(status='Present').count()
    absent  = attendances.filter(status='Absent').count()
    late    = attendances.filter(status='Late').count()
    on_leave= attendances.filter(status='Leave').count()

    context = {
        'attendances': attendances,
        'date_from': date_from,
        'date_to': date_to,
        'total': total, 'present': present,
        'absent': absent, 'late': late, 'on_leave': on_leave,
    }
    return render(request, 'hrm/attendance_report.html', context)


def attendance_pdf_view(request):
    """Generate and download a PDF of the attendance report using a BytesIO buffer."""
    from datetime import date, datetime
    from io import BytesIO
    from xhtml2pdf import pisa
    from django.template.loader import get_template

    today = date.today()
    date_from_str = request.GET.get('date_from', today.replace(day=1).strftime('%Y-%m-%d'))
    date_to_str   = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        date_from = today.replace(day=1)
        date_to   = today

    attendances = (
        Attendance.objects
        .filter(date__range=[date_from, date_to])
        .select_related('employee', 'employee__department', 'employee__designation')
        .order_by('date', 'employee__first_name')
    )

    total    = attendances.count()
    present  = attendances.filter(status='Present').count()
    absent   = attendances.filter(status='Absent').count()
    late     = attendances.filter(status='Late').count()
    on_leave = attendances.filter(status='Leave').count()

    context = {
        'attendances': attendances,
        'date_from': date_from,
        'date_to': date_to,
        'total': total, 'present': present,
        'absent': absent, 'late': late, 'on_leave': on_leave,
        'generated_at': datetime.now().strftime('%d %B %Y, %I:%M %p'),
    }

    template    = get_template('hrm/attendance_pdf.html')
    html_string = template.render(context, request)

    # Render to a BytesIO buffer — much more reliable than writing to response directly
    buffer = BytesIO()
    pisa_status = pisa.CreatePDF(html_string, dest=buffer)

    if pisa_status.err:
        return HttpResponse(f"PDF generation error: {pisa_status.err}", status=500)

    filename = f"attendance_report_{date_from}_to_{date_to}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def attendance_excel_view(request):
    """Generate and download an Excel spreadsheet of the attendance report."""
    from datetime import date, datetime
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = date.today()
    date_from_str = request.GET.get('date_from', today.replace(day=1).strftime('%Y-%m-%d'))
    date_to_str   = request.GET.get('date_to', today.strftime('%Y-%m-%d'))

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        date_from = today.replace(day=1)
        date_to   = today

    attendances = (
        Attendance.objects
        .filter(date__range=[date_from, date_to])
        .select_related('employee', 'employee__department', 'employee__designation')
        .order_by('date', 'employee__first_name')
    )

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    header_fill   = PatternFill("solid", fgColor="1D4ED8")   # Blue
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    summary_fill  = PatternFill("solid", fgColor="EFF6FF")
    title_font    = Font(bold=True, size=14, color="1D4ED8")
    normal_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0"),
    )

    status_colors = {
        'Present': ("DCFCE7", "166534"),
        'Late':    ("FEF9C3", "854D0E"),
        'Absent':  ("FEE2E2", "991B1B"),
        'Leave':   ("DBEAFE", "1E40AF"),
    }

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "Plasma IT Solutions Ltd. — Attendance Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Period: {date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}  |  Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(color="64748B", size=10)
    ws.row_dimensions[2].height = 18

    # Summary row
    total    = attendances.count()
    present  = attendances.filter(status='Present').count()
    late     = attendances.filter(status='Late').count()
    absent   = attendances.filter(status='Absent').count()
    on_leave = attendances.filter(status='Leave').count()

    ws['A4'] = "Total Records"
    ws['B4'] = total
    ws['C4'] = "Present"
    ws['D4'] = present
    ws['E4'] = "Late"
    ws['F4'] = late
    ws['G4'] = "Absent"
    ws['H4'] = absent
    ws['I4'] = "On Leave"
    ws['J4'] = on_leave

    for cell in ws[4]:
        cell.fill = summary_fill
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')

    # Header row
    headers = ['#', 'Date', 'Employee Name', 'Employee Code', 'Department',
               'Designation', 'Check In', 'Check Out', 'Working Hours', 'Status']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = normal_border
    ws.row_dimensions[6].height = 20

    # Data rows
    for row_idx, record in enumerate(attendances, start=1):
        row_num = row_idx + 6
        check_in_val  = record.check_in.strftime('%H:%M')  if record.check_in  else '-'
        check_out_val = record.check_out.strftime('%H:%M') if record.check_out else '-'

        row_data = [
            row_idx,
            record.date.strftime('%d %b %Y'),
            f"{record.employee.first_name} {record.employee.last_name}",
            record.employee.employee_code,
            record.employee.department.name if record.employee.department else '-',
            record.employee.designation.name if record.employee.designation else '-',
            check_in_val,
            check_out_val,
            float(record.working_hours) if record.working_hours else '-',
            record.status,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.border = normal_border
            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

        # Color Status cell
        status_cell = ws.cell(row=row_num, column=10)
        if record.status in status_colors:
            bg, fg = status_colors[record.status]
            status_cell.fill = PatternFill("solid", fgColor=bg)
            status_cell.font = Font(bold=True, color=fg)

    # Column widths
    col_widths = [5, 14, 22, 16, 28, 28, 10, 10, 14, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header
    ws.freeze_panes = 'A7'

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"attendance_report_{date_from}_to_{date_to}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response